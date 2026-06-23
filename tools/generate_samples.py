"""One-shot script to generate the demo PDF and Excel samples.

Run manually from the repo root::

    .\\.venv\\Scripts\\python tools\\generate_samples.py

The output files (``samples/knowledge_assist_policy.pdf`` and
``samples/launch_schedule.xlsx``) are committed to the repo so reviewers do
not need to install ``fpdf2``. ``fpdf2`` is only required to regenerate the
PDF; it is intentionally NOT listed in ``requirements.txt``.
"""
from __future__ import annotations

from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


SAMPLES_DIR = Path(__file__).resolve().parent.parent / "samples"


def build_policy_pdf() -> None:
    pdf = FPDF()
    pdf.set_margins(left=20, top=20, right=20)
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    body_width = pdf.epw  # effective page width (page minus left+right margins)

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(body_width, 10, "Knowledge Assist pilot - policy overview")
    pdf.ln(12)

    pdf.set_font("Helvetica", "", 11)
    intro = (
        "This policy applies to the Knowledge Assist pilot launching with the "
        "customer support and finance operations teams. It describes which document "
        "sources are approved, how reviewers should evaluate answers, and how to "
        "escalate when an answer cannot be trusted."
    )
    pdf.multi_cell(body_width, 6, intro)
    pdf.ln(4)

    sections = [
        (
            "1. Approved document sources",
            "Only documents owned by the Policy team and explicitly tagged as pilot-approved "
            "may be uploaded into the assistant. Personal notes, draft contracts, and any "
            "document containing customer personal data are out of scope for the pilot.",
        ),
        (
            "2. Reviewer responsibilities",
            "Every answer must be reviewed by a human before it is shared with a customer. "
            "Reviewers confirm that the cited sources actually support the answer, that "
            "weak-retrieval warnings are taken seriously, and that the assistant is not "
            "asked to make policy decisions on its own.",
        ),
        (
            "3. Escalation procedure",
            "If an answer disagrees with an approved policy document, or if the assistant "
            "returns a retrieval-only answer because the LLM was unavailable, the reviewer "
            "must escalate to the Digital Workplace team within one business day. The "
            "escalation note should include the question, the answer, and the cited sources.",
        ),
        (
            "4. Known limitations",
            "Scanned PDFs without extractable text are rejected during upload. Large "
            "spreadsheets are converted into row-oriented text, which can lose layout "
            "context. Answers cite source numbers, but the reviewer remains responsible "
            "for verifying that the cited passage actually contains the cited fact.",
        ),
        (
            "5. Owners",
            "Policy approval: Policy team. Reviewer training: Maya. Demo walkthrough and "
            "release notes: Omar. Production rollout decision: Digital Workplace lead.",
        ),
    ]

    for heading, body in sections:
        pdf.set_font("Helvetica", "B", 12)
        pdf.multi_cell(body_width, 7, heading)
        pdf.set_font("Helvetica", "", 11)
        pdf.multi_cell(body_width, 6, body)
        pdf.ln(3)

    pdf.output(str(SAMPLES_DIR / "knowledge_assist_policy.pdf"))


def build_schedule_xlsx() -> None:
    workbook = Workbook()

    milestones = workbook.active
    milestones.title = "Milestones"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    headers = ["Phase", "Start", "End", "Owner", "Status", "Notes"]
    milestones.append(headers)
    for cell in milestones[1]:
        cell.font = header_font
        cell.fill = header_fill

    milestone_rows = [
        ("Discovery", "2026-05-15", "2026-05-29", "Digital Workplace", "Done", "Confirmed pilot scope with support and finance ops."),
        ("Document approval", "2026-05-30", "2026-07-03", "Policy team", "At risk", "Approved set incomplete; this is the highest launch risk."),
        ("Indexing dry run", "2026-06-05", "2026-06-20", "Maya", "On track", "Validates ingestion of PDF, TXT, CSV, Excel uploads."),
        ("Reviewer training", "2026-06-22", "2026-07-08", "Maya", "On track", "Reviewers learn how to inspect retrieved vs used sources."),
        ("Demo walkthrough", "2026-07-09", "2026-07-12", "Omar", "On track", "Captures the same flow used in the README demo checklist."),
        ("Pilot launch", "2026-07-15", "2026-07-15", "Digital Workplace lead", "Pending", "Go/no-go depends on the policy approval milestone."),
    ]
    for row in milestone_rows:
        milestones.append(row)

    for column_index, width in enumerate([20, 14, 14, 22, 12, 60], start=1):
        milestones.column_dimensions[milestones.cell(row=1, column=column_index).column_letter].width = width

    dependencies = workbook.create_sheet("Dependencies")
    dependencies.append(["Item", "Depends on", "Status", "Owner"])
    for cell in dependencies[1]:
        cell.font = header_font
        cell.fill = header_fill
    dependency_rows = [
        ("Pilot launch", "Document approval", "At risk", "Policy team"),
        ("Reviewer training", "Indexing dry run", "On track", "Maya"),
        ("Demo walkthrough", "Reviewer training", "On track", "Omar"),
        ("Production rollout decision", "Pilot launch", "Pending", "Digital Workplace lead"),
    ]
    for row in dependency_rows:
        dependencies.append(row)
    for column_index, width in enumerate([28, 28, 12, 22], start=1):
        dependencies.column_dimensions[dependencies.cell(row=1, column=column_index).column_letter].width = width

    workbook.save(SAMPLES_DIR / "launch_schedule.xlsx")


def main() -> None:
    SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
    build_policy_pdf()
    build_schedule_xlsx()
    print(f"Wrote samples to {SAMPLES_DIR}")


if __name__ == "__main__":
    main()
